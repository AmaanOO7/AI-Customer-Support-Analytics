"""
Export Handler Module
Handles exporting data to various formats (Excel, CSV, JSON)
"""

import os
import json
import logging
from typing import Dict, Any, List
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ExportHandler:
    """Handles data export operations"""
    
    def __init__(self, export_folder: str = 'exports'):
        """
        Initialize ExportHandler
        
        Args:
            export_folder: Directory for exported files
        """
        self.export_folder = export_folder
        os.makedirs(export_folder, exist_ok=True)
    
    def export_to_csv(self, data: List[Dict[str, Any]], output_path: str) -> str:
        """
        Export data to CSV format
        
        Args:
            data: List of dictionaries containing data
            output_path: Path to save CSV file
            
        Returns:
            Path to exported file
        """
        try:
            df = pd.DataFrame(data)
            df.to_csv(output_path, index=False, encoding='utf-8')
            logger.info(f"Data exported to CSV: {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"Error exporting to CSV: {str(e)}")
            raise
    
    def export_to_json(self, data: Any, output_path: str, indent: int = 2) -> str:
        """
        Export data to JSON format
        
        Args:
            data: Data to export (dict or list)
            output_path: Path to save JSON file
            indent: JSON indentation level
            
        Returns:
            Path to exported file
        """
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=indent, ensure_ascii=False, default=str)
            logger.info(f"Data exported to JSON: {output_path}")
            return output_path
        except Exception as e:
            logger.error(f"Error exporting to JSON: {str(e)}")
            raise
    
    def export_to_excel(self, data: Dict[str, Any], output_path: str) -> str:
        """
        Export comprehensive analytics to Excel with multiple sheets
        
        Args:
            data: Dictionary containing all analytics data
            output_path: Path to save Excel file
            
        Returns:
            Path to exported file
        """
        try:
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Create sheets
            self._create_summary_sheet(wb, data)
            self._create_tickets_sheet(wb, data)
            self._create_analytics_sheet(wb, data)
            self._create_insights_sheet(wb, data)
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"Data exported to Excel: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}")
            raise
    
    def _create_summary_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create summary sheet with KPIs"""
        ws = wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = "Customer Support Analytics Summary"
        ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:D1')
        ws.row_dimensions[1].height = 30
        
        # Get summary data
        summary = data.get('analytics_summary', {})
        
        # KPIs
        kpis = [
            ['Metric', 'Value'],
            ['Total Tickets', summary.get('total_tickets', 0)],
            ['Total Customers', summary.get('total_customers', 0)],
            ['Average Customer Satisfaction', f"{summary.get('avg_customer_satisfaction', 0):.2f}%"],
            ['Average Agent Score', f"{summary.get('avg_agent_score', 0):.2f}%"],
            ['Positive Tickets', summary.get('positive_tickets', 0)],
            ['Negative Tickets', summary.get('negative_tickets', 0)],
            ['Critical Tickets', summary.get('critical_tickets', 0)],
            ['Escalated Tickets', summary.get('escalation_stats', {}).get('yes', 0)],
            ['Escalation Rate', f"{summary.get('escalation_stats', {}).get('percentage', 0):.2f}%"]
        ]
        
        # Write KPIs
        for row_idx, row_data in enumerate(kpis, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:  # Header row
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                else:
                    if col_idx == 1:
                        cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='left' if col_idx == 1 else 'center')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        
        # Apply borders
        self._apply_borders(ws, 3, len(kpis) + 2, 1, 2)
    
    def _create_tickets_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create detailed tickets sheet"""
        ws = wb.create_sheet("Tickets")
        
        # Get analysis results
        analysis_results = data.get('analysis_results', [])
        
        if not analysis_results:
            ws['A1'] = "No ticket data available"
            return
        
        # Prepare data for export
        export_data = []
        for result in analysis_results:
            export_data.append({
                'Ticket ID': result.get('ticket_id', ''),
                'Customer Name': result.get('customer_name', ''),
                'Agent Name': result.get('agent_name', ''),
                'Date': result.get('date', ''),
                'Product': result.get('product', ''),
                'Sentiment': result.get('sentiment', ''),
                'Emotion': result.get('emotion', ''),
                'Intent': result.get('intent', ''),
                'Issue Category': result.get('issue_category', ''),
                'Priority': result.get('priority', ''),
                'Urgency': result.get('urgency', ''),
                'Satisfaction Score': result.get('customer_satisfaction_score', ''),
                'Churn Risk': result.get('churn_risk', ''),
                'Escalation Needed': result.get('escalation_needed', ''),
                'Customer Message': result.get('customer_message', '')[:100] + '...' if len(result.get('customer_message', '')) > 100 else result.get('customer_message', ''),
                'Complaint Summary': result.get('complaint_summary', '')
            })
        
        # Convert to DataFrame and write
        df = pd.DataFrame(export_data)
        
        # Write headers
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data
        for row_idx, row_data in enumerate(df.values, start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Adjust column widths
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.column_dimensions[chr(64 + col_idx)].width = 15
        
        # Apply borders
        self._apply_borders(ws, 1, len(df) + 1, 1, len(df.columns))
    
    def _create_analytics_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create analytics distributions sheet"""
        ws = wb.create_sheet("Analytics")
        
        summary = data.get('analytics_summary', {})
        
        row = 1
        
        # Sentiment Distribution
        ws.cell(row=row, column=1, value="Sentiment Distribution")
        ws.cell(row=row, column=1).font = Font(size=14, bold=True, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        sentiment_dist = summary.get('sentiment_distribution', {})
        ws.cell(row=row, column=1, value="Sentiment")
        ws.cell(row=row, column=2, value="Count")
        ws.cell(row=row, column=3, value="Percentage")
        self._style_header_row(ws, row, 1, 3)
        row += 1
        
        total = sum(sentiment_dist.values()) if sentiment_dist else 1
        for sentiment, count in sorted(sentiment_dist.items(), key=lambda x: x[1], reverse=True):
            ws.cell(row=row, column=1, value=sentiment)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{(count/total*100):.2f}%")
            row += 1
        
        row += 2
        
        # Category Distribution
        ws.cell(row=row, column=1, value="Issue Category Distribution")
        ws.cell(row=row, column=1).font = Font(size=14, bold=True, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        category_dist = summary.get('category_distribution', {})
        ws.cell(row=row, column=1, value="Category")
        ws.cell(row=row, column=2, value="Count")
        ws.cell(row=row, column=3, value="Percentage")
        self._style_header_row(ws, row, 1, 3)
        row += 1
        
        total = sum(category_dist.values()) if category_dist else 1
        for category, count in sorted(category_dist.items(), key=lambda x: x[1], reverse=True):
            ws.cell(row=row, column=1, value=category)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{(count/total*100):.2f}%")
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def _create_insights_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create business insights sheet"""
        ws = wb.create_sheet("Business Insights")
        
        insights = data.get('business_insights', {})
        
        row = 1
        
        # Executive Summary
        ws.cell(row=row, column=1, value="Executive Summary")
        ws.cell(row=row, column=1).font = Font(size=14, bold=True, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        exec_summary = insights.get('executive_summary', 'No summary available')
        ws.cell(row=row, column=1, value=exec_summary)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(f'A{row}:D{row}')
        ws.row_dimensions[row].height = 60
        row += 3
        
        # Recommendations
        ws.cell(row=row, column=1, value="Actionable Recommendations")
        ws.cell(row=row, column=1).font = Font(size=14, bold=True, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        recommendations = insights.get('actionable_recommendations', [])
        for i, rec in enumerate(recommendations, 1):
            ws.cell(row=row, column=1, value=f"{i}.")
            ws.cell(row=row, column=2, value=rec)
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(f'B{row}:D{row}')
            ws.row_dimensions[row].height = 30
            row += 1
        
        row += 2
        
        # Business Risks
        ws.cell(row=row, column=1, value="Business Risks")
        ws.cell(row=row, column=1).font = Font(size=14, bold=True, color="FFFFFF")
        ws.cell(row=row, column=1).fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        risks = insights.get('business_risks', [])
        for i, risk in enumerate(risks, 1):
            ws.cell(row=row, column=1, value=f"{i}.")
            ws.cell(row=row, column=2, value=risk)
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(f'B{row}:D{row}')
            ws.row_dimensions[row].height = 30
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def _style_header_row(self, ws, row: int, start_col: int, end_col: int):
        """Apply header styling to a row"""
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
    
    def _apply_borders(self, ws, start_row: int, end_row: int, start_col: int, end_col: int):
        """Apply borders to a range of cells"""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                ws.cell(row=row, column=col).border = thin_border
